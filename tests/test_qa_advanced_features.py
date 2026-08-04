import pytest
import os
import json
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app.main import app
from app.adapters.memory.memory import _get_connection, memory
from app.domain.services.tax_parser_service import TaxParserService
from app.domain.services.boe_reader import BOEReaderService
from app.core.websocket_manager import guardian_ws_manager

@pytest.fixture(autouse=True)
def clean_db_and_excel():
    """Limpia el estado de la base de datos de test antes de cada ejecución."""
    with _get_connection() as conn:
        conn.execute("DROP TABLE IF EXISTS invoices")
        conn.commit()
        
        # Recrear tabla de facturas
        from app.adapters.memory.memory import _init_db_schema
        _init_db_schema(conn)
        
    memory._cache.clear()
    
    excel_path = "data/facturas_alfonso.xlsx"
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except OSError:
            pass
            
    yield
    
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)
        except OSError:
            pass

# ── 1. TEST INTEGRACIÓN & QA: EXCEL AUTOMÁTICO ──────────────────────────────────
def test_invoice_saving_triggers_excel_sync():
    """
    Verifica que al guardar una factura:
    1. Se guarde cifrada en SQLite.
    2. Se dispare automáticamente la sincronización y genere el Excel en disco descifrado.
    3. El archivo Excel tenga las pestañas 'Ingresos' y 'Gastos' bien formateadas.
    """
    invoice_data = {
        "invoice_id": "FAC-2026-001",
        "date": "03/08/2026",
        "issuer_name": "Servicios Digitales SL",
        "issuer_nif": "B12345678",
        "receiver_name": "Luis Domingo Pérez",
        "receiver_nif": "12345678Z",
        "base_imponible": 500.0,
        "iva_rate": 21.0,
        "iva_amount": 105.0,
        "irpf_rate": 15.0,
        "irpf_amount": 75.0,
        "total_amount": 530.0,
        "category": "gasto",
        "quarter": 3,
        "year": 2026
    }

    # Guardar factura (debe disparar el trigger de ExcelSyncService)
    last_id = TaxParserService.save_invoice_to_db(invoice_data)
    assert last_id > 0

    excel_path = "data/facturas_alfonso.xlsx"
    assert os.path.exists(excel_path), "El archivo Excel local no se creó tras el trigger."

    # Cargar y verificar el Excel
    wb = load_workbook(excel_path)
    assert "Ingresos" in wb.sheetnames
    assert "Gastos" in wb.sheetnames

    # El registro es un "gasto", por lo que debe estar en la pestaña Gastos
    ws_gastos = wb["Gastos"]
    assert ws_gastos.max_row == 2  # Fila 1: Headers, Fila 2: Datos

    row_values = [cell.value for cell in ws_gastos[2]]
    assert row_values[0] == "FAC-2026-001"  # ID Factura (Descifrada)
    assert row_values[2] == "Servicios Digitales SL"  # Emisor (Descifrada)
    assert row_values[6] == 500.0  # Base Imponible (Float)
    assert row_values[11] == 530.0  # Total (Float)

# ── 2. TEST INTEGRACIÓN & QA: WEBSOCKET GUARDIÁN ──────────────────────────────
def test_websocket_guardian_communication():
    """
    Verifica que la extensión de Chrome pueda establecer conexión WebSocket,
    enviar un mensaje de confirmación y que el backend pueda retransmitir alertas.
    """
    client = TestClient(app)
    
    with client.websocket_connect("/ws/guardian") as websocket:
        # Enviar un mensaje de prueba desde la extensión simulada
        test_payload = {"type": "handshake", "client_id": "test_ext"}
        websocket.send_json(test_payload)
        
        # Verificar que el WebSocket recibe el eco de confirmación del servidor
        response = websocket.receive_json()
        assert response["status"] == "received"
        assert response["echo"]["client_id"] == "test_ext"

# ── 3. TEST INTEGRACIÓN & QA: MOCK BOE READER ─────────────────────────────────
@pytest.mark.asyncio
async def test_boe_fetch_and_analysis_mock(monkeypatch):
    """
    Simula la descarga de un sumario XML del BOE con decretos de prueba
    y verifica que se filtren e interpreten correctamente las alertas fiscales.
    """
    mock_xml_content = """<?xml version="1.0" encoding="utf-8"?>
    <sumario>
      <diario id="BOE-A-2026-1234">
        <seccion num="1">
          <item id="BOE-A-2026-001">
            <titulo>Real Decreto-ley de medidas urgentes en materia de IVA de alimentos.</titulo>
            <link>https://www.boe.es/boe/dias/2026/08/03/pdfs/BOE-A-2026-001.pdf</link>
          </item>
          <item id="BOE-A-2026-002">
            <titulo>Nombramientos en la Administración del Estado.</titulo>
            <link>https://www.boe.es/boe/dias/2026/08/03/pdfs/BOE-A-2026-002.pdf</link>
          </item>
          <item id="BOE-A-2026-003">
            <titulo>Resolución sobre las nuevas bases de cotización de autónomos.</titulo>
            <link>https://www.boe.es/boe/dias/2026/08/03/pdfs/BOE-A-2026-003.pdf</link>
          </item>
        </seccion>
      </diario>
    </sumario>
    """.encode("utf-8")

    class MockResponse:
        def __init__(self, content, status_code=200):
            self.content = content
            self.status_code = status_code

    async def mock_get(*args, **kwargs):
        return MockResponse(mock_xml_content)

    import httpx
    monkeypatch.setattr(httpx.AsyncClient, "get", mock_get)

    # Ejecutar ingesta simulada
    alerts = await BOEReaderService.fetch_and_parse_boe("20260803")
    
    # Debería capturar solo el Decreto de IVA (001) y la Cotización de Autónomos (003). 
    # El Nombramiento (002) no coincide con ninguna palabra clave.
    assert len(alerts) == 2
    
    alert_ids = [a["id"] for a in alerts]
    assert "BOE-A-2026-001" in alert_ids
    assert "BOE-A-2026-003" in alert_ids
    assert "BOE-A-2026-002" not in alert_ids

    # Ejecutar análisis
    suggested = await BOEReaderService.analyze_fiscal_alerts(alerts)
    assert len(suggested) == 2
    
    types = [s["tipo"] for s in suggested]
    assert "IVA" in types
    assert "Seguridad Social" in types
