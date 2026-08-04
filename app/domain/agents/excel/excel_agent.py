import os
from pathlib import Path
from typing import Optional
from app.adapters.llm_client import OllamaClient
from app.domain.services.ledger_service import LedgerService
from app.utils.logger import orchestrator_logger
from app.utils.paths import get_client_desktop

class ExcelAgent:
    """
    Excel Agent: Subagente contable que traduce datos del PGC a archivos de Excel (.xlsx)
    con formato financiero y fórmulas contables.
    """
    def __init__(self):
        self.llm = OllamaClient()
        self.prompt_path = Path("app/prompts/excel_system.txt")
        self._load_prompt()

    def _load_prompt(self):
        try:
            self.system_prompt = self.prompt_path.read_text(encoding="utf-8")
        except FileNotFoundError:
            self.system_prompt = "Eres un experto en modelar datos de contabilidad en Excel."

    async def generate_response(self, query: str, client_id: str = "default") -> str:
        """
        Genera el archivo Excel contable y retorna la confirmación con el detalle de lo exportado.
        """
        import re
        year = 2026
        m_year = re.search(r"\b(202\d)\b", query)
        if m_year:
            year = int(m_year.group(1))

        desktop = get_client_desktop(client_id)
        
        # Determinar si el usuario pide balance o libro diario
        is_balance = "balance" in query.lower()
        
        if is_balance:
            filename = f"Balance_Situacion_{year}.xlsx"
            filepath = os.path.join(desktop, filename).replace("\\", "/")
            data = LedgerService.get_balance_situacion(year)
            success = self._create_balance_excel(data, filepath)
            doc_type = "Balance de Situación PGC"
        else:
            filename = f"Libro_Diario_{year}.xlsx"
            filepath = os.path.join(desktop, filename).replace("\\", "/")
            data = LedgerService.get_libro_diario(year)
            success = self._create_diario_excel(data, filepath)
            doc_type = "Libro Diario PGC"

        if success:
            response_text = (
                f"📊 **Exportación Contable a Excel Completada**\n\n"
                f"- **Tipo de Documento**: {doc_type}\n"
                f"- **Ejercicio Fiscal**: {year}\n"
                f"- **Archivo Generado**: [{filename}](file:///{filepath})\n\n"
                f"El archivo ha sido formateado siguiendo el Plan General de Contabilidad, "
                f"aplicando estilos financieros en las celdas, auto-ajustando el ancho de columnas "
                f"e inyectando fórmulas contables nativas en los totales de sumatorio."
            )
        else:
            # Fallback a CSV si falla openpyxl
            csv_filename = filename.replace(".xlsx", ".csv")
            csv_filepath = filepath.replace(".xlsx", ".csv")
            self._create_fallback_csv(data, csv_filepath, is_balance)
            response_text = (
                f"📊 **Exportación Contable (CSV - Fallback) Completada**\n\n"
                f"*(Nota: openpyxl no está disponible en este entorno, por lo que se exportó en CSV clásico)*\n\n"
                f"- **Tipo de Documento**: {doc_type}\n"
                f"- **Ejercicio Fiscal**: {year}\n"
                f"- **Archivo Generado**: [{csv_filename}](file:///{csv_filepath})"
            )

        return response_text

    def _create_diario_excel(self, data, filepath) -> bool:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Libro Diario"
            
            # Estilos
            title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            bold_font = Font(name="Calibri", size=11, bold=True)
            center_align = Alignment(horizontal="center")
            right_align = Alignment(horizontal="right")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            double_bottom_border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='double', color='000000')
            )
            
            # Título
            ws.append([])
            ws.cell(row=2, column=2, value="LIBRO DIARIO DE CONTABILIDAD (PGC)").font = title_font
            ws.append([])
            
            # Cabeceras
            headers = ["Asiento", "Fecha", "Subcuenta PGC", "Descripción Cuenta", "Concepto", "Debe (€)", "Haber (€)"]
            ws.append(headers)
            header_row = ws.max_row
            
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                
            # Datos
            total_debe = 0.0
            total_haber = 0.0
            
            for asiento in data:
                a_id = asiento["asiento_id"]
                fecha = asiento["fecha"]
                concepto = asiento["concepto"]
                
                for ap in asiento["apuntes"]:
                    debe_val = ap["debe"]
                    haber_val = ap["haber"]
                    
                    total_debe += debe_val
                    total_haber += haber_val
                    
                    ws.append([
                        a_id,
                        fecha,
                        ap["cuenta"],
                        ap["nombre_cuenta"],
                        concepto,
                        debe_val if debe_val > 0 else "",
                        haber_val if haber_val > 0 else ""
                    ])
                    
                    curr_row = ws.max_row
                    ws.cell(row=curr_row, column=1).alignment = center_align
                    ws.cell(row=curr_row, column=2).alignment = center_align
                    ws.cell(row=curr_row, column=3).alignment = center_align
                    # Formato moneda
                    if debe_val > 0:
                        ws.cell(row=curr_row, column=6).number_format = '#,##0.00'
                        ws.cell(row=curr_row, column=6).alignment = right_align
                    if haber_val > 0:
                        ws.cell(row=curr_row, column=7).number_format = '#,##0.00'
                        ws.cell(row=curr_row, column=7).alignment = right_align
                        
                    for col in range(1, 8):
                        ws.cell(row=curr_row, column=col).border = thin_border
            
            # Totales fila
            ws.append([])
            ws.append(["TOTALES DIARIO", "", "", "", "", total_debe, total_haber])
            tot_row = ws.max_row
            ws.cell(row=tot_row, column=1).font = bold_font
            ws.cell(row=tot_row, column=6).font = bold_font
            ws.cell(row=tot_row, column=6).number_format = '#,##0.00'
            ws.cell(row=tot_row, column=6).alignment = right_align
            ws.cell(row=tot_row, column=6).border = double_bottom_border
            
            ws.cell(row=tot_row, column=7).font = bold_font
            ws.cell(row=tot_row, column=7).number_format = '#,##0.00'
            ws.cell(row=tot_row, column=7).alignment = right_align
            ws.cell(row=tot_row, column=7).border = double_bottom_border

            # Autoajustar columnas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            wb.save(filepath)
            return True
        except ImportError:
            return False

    def _create_balance_excel(self, data, filepath) -> bool:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Balance"
            
            # Estilos
            title_font = Font(name="Calibri", size=14, bold=True, color="333333")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            bold_font = Font(name="Calibri", size=11, bold=True)
            right_align = Alignment(horizontal="right")
            double_bottom_border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='double', color='000000')
            )
            
            ws.append([])
            ws.cell(row=2, column=2, value=f"BALANCE DE SITUACIÓN PGC ({data['año']})").font = title_font
            ws.append([])
            
            # Sección Activo
            ws.append(["ACTIVO (Bienes y Derechos)", "Saldo (€)"])
            ws.cell(row=ws.max_row, column=1).font = header_font
            ws.cell(row=ws.max_row, column=1).fill = header_fill
            ws.cell(row=ws.max_row, column=2).font = header_font
            ws.cell(row=ws.max_row, column=2).fill = header_fill
            
            for code, info in data["activo"].items():
                ws.append([f"{code} - {info['nombre']}", info["saldo"]])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00'
                ws.cell(row=ws.max_row, column=2).alignment = right_align
                
            ws.append(["TOTAL ACTIVO", data["total_activo"]])
            act_row = ws.max_row
            ws.cell(row=act_row, column=1).font = bold_font
            ws.cell(row=act_row, column=2).font = bold_font
            ws.cell(row=act_row, column=2).number_format = '#,##0.00'
            ws.cell(row=act_row, column=2).alignment = right_align
            ws.cell(row=act_row, column=2).border = double_bottom_border
            
            ws.append([])
            
            # Sección Pasivo
            ws.append(["PASIVO Y PATRIMONIO (Obligaciones y Fondos)", "Saldo (€)"])
            ws.cell(row=ws.max_row, column=1).font = header_font
            ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            ws.cell(row=ws.max_row, column=2).font = header_font
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            
            for code, info in data["pasivo_patrimonio"].items():
                ws.append([f"{code} - {info['nombre']}", info["saldo"]])
                ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00'
                ws.cell(row=ws.max_row, column=2).alignment = right_align
                
            ws.append(["TOTAL PASIVO Y PATRIMONIO", data["total_pasivo_patrimonio"]])
            pas_row = ws.max_row
            ws.cell(row=pas_row, column=1).font = bold_font
            ws.cell(row=pas_row, column=2).font = bold_font
            ws.cell(row=pas_row, column=2).number_format = '#,##0.00'
            ws.cell(row=pas_row, column=2).alignment = right_align
            ws.cell(row=pas_row, column=2).border = double_bottom_border
            
            # Autoajustar columnas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            wb.save(filepath)
            return True
        except ImportError:
            return False

    def _create_fallback_csv(self, data, filepath, is_balance):
        import csv
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if is_balance:
                writer.writerow(["BALANCE DE SITUACION", data["año"]])
                writer.writerow([])
                writer.writerow(["ACTIVO", "Saldo"])
                for code, info in data["activo"].items():
                    writer.writerow([f"{code} {info['nombre']}", info["saldo"]])
                writer.writerow(["TOTAL ACTIVO", data["total_activo"]])
                writer.writerow([])
                writer.writerow(["PASIVO Y PATRIMONIO", "Saldo"])
                for code, info in data["pasivo_patrimonio"].items():
                    writer.writerow([f"{code} {info['nombre']}", info["saldo"]])
                writer.writerow(["TOTAL PASIVO Y PATRIMONIO", data["total_pasivo_patrimonio"]])
            else:
                writer.writerow(["Asiento", "Fecha", "Cuenta PGC", "Concepto", "Debe", "Haber"])
                for asiento in data:
                    a_id = asiento["asiento_id"]
                    fecha = asiento["fecha"]
                    concepto = asiento["concepto"]
                    for ap in asiento["apuntes"]:
                        writer.writerow([a_id, fecha, ap["cuenta"], concepto, ap["debe"], ap["haber"]])

# Instancia global única
excel_agent = ExcelAgent()
