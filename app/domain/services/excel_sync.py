import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.adapters.memory.memory import _get_connection
from app.utils.encryption import encryptor

logger = logging.getLogger("excel_sync")

class ExcelSyncService:
    """
    Servicio para exportar y sincronizar facturas desde la base de datos local SQLite
    a un archivo Excel (.xlsx) local de forma inalterable y local-first.
    """

    @classmethod
    def sync_invoices_to_excel(cls, output_path: str = "data/facturas_alfonso.xlsx") -> str:
        """
        Lee todas las facturas de la base de datos, las desencripta y las escribe
        en un libro de Excel estructurado con pestañas independientes para Ingresos y Gastos.
        """
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        # Eliminar pestaña por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crear pestañas para ingresos y gastos
        ws_ingresos = wb.create_sheet(title="Ingresos")
        ws_gastos = wb.create_sheet(title="Gastos")

        headers = [
            "ID Factura", "Fecha", "Emisor", "NIF Emisor", "Receptor", "NIF Receptor",
            "Base Imponible (€)", "IVA (%)", "Cuota IVA (€)", "IRPF (%)", "Cuota IRPF (€)",
            "Total (€)", "Trimestre", "Año"
        ]

        # Estilos profesionales (Azul para headers de Ingresos, Gris/Negro para Gastos)
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_ingresos = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Azul
        fill_gastos = PatternFill(start_color="404040", end_color="404040", fill_type="solid")    # Gris Oscuro
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for ws, fill in [(ws_ingresos, fill_ingresos), (ws_gastos, fill_gastos)]:
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = fill
                cell.alignment = align_center

        # Leer de la base de datos
        with _get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT invoice_id, date, issuer_name, issuer_nif, receiver_name, receiver_nif,
                       base_imponible, iva_rate, iva_amount, irpf_rate, irpf_amount, total_amount,
                       category, quarter, year 
                FROM invoices ORDER BY year DESC, quarter DESC, id DESC
            """)
            rows = cursor.fetchall()

        count_ingresos = 0
        count_gastos = 0

        for row in rows:
            try:
                # Desencriptar datos
                invoice_id = encryptor.decrypt(row["invoice_id"])
                date_str = encryptor.decrypt(row["date"])
                issuer_name = encryptor.decrypt(row["issuer_name"])
                issuer_nif = encryptor.decrypt(row["issuer_nif"])
                receiver_name = encryptor.decrypt(row["receiver_name"])
                receiver_nif = encryptor.decrypt(row["receiver_nif"])
                base_imponible = float(encryptor.decrypt(row["base_imponible"]))
                iva_rate = float(encryptor.decrypt(row["iva_rate"]))
                iva_amount = float(encryptor.decrypt(row["iva_amount"]))
                irpf_rate = float(encryptor.decrypt(row["irpf_rate"]))
                irpf_amount = float(encryptor.decrypt(row["irpf_amount"]))
                total_amount = float(encryptor.decrypt(row["total_amount"]))
                category = row["category"]
                quarter = int(row["quarter"])
                year = int(row["year"])

                row_data = [
                    invoice_id, date_str, issuer_name, issuer_nif, receiver_name, receiver_nif,
                    base_imponible, iva_rate, iva_amount, irpf_rate, irpf_amount,
                    total_amount, quarter, year
                ]

                # Determinar pestaña destino
                if category.lower() in ("ingreso", "income"):
                    ws = ws_ingresos
                    count_ingresos += 1
                else:
                    ws = ws_gastos
                    count_gastos += 1

                ws.append(row_data)

                # Aplicar alineación y formatos
                current_row = ws.max_row
                # IDs, Fechas, NIFs y Trimestre/Año centrados
                for col in (1, 2, 4, 6, 13, 14):
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = align_center
                    cell.border = border_thin
                
                # Nombres a la izquierda con bordes
                for col in (3, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = border_thin

                # Valores monetarios con formato número y alineación derecha
                for col in (7, 8, 9, 10, 11, 12):
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = align_right
                    cell.number_format = '0.00'
                    cell.border = border_thin

            except Exception as e:
                logger.error(f"Error procesando fila de factura para Excel: {e}")
                continue

        # Autoajustar ancho de columnas para legibilidad
        for ws in [ws_ingresos, ws_gastos]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        logger.info(f"Excel sincronizado correctamente en {output_path}. Ingresos: {count_ingresos}, Gastos: {count_gastos}")
        return output_path
